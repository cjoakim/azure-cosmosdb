__author__  = 'Chris Joakim'
__email__   = "chjoakim@microsoft.com,christopher.joakim@gmail.com"
__license__ = "MIT"
__version__ = "2020.05.28"

import csv
import json
import os

import arrow 

from openpyxl import Workbook
from openpyxl.styles import Alignment, Color, Font, Fill
from openpyxl.styles import colors

class Excel(object):

    def __init__(self):
        pass

    def generate(self, title, items, attributes, outfile):
        wb = Workbook()
        ws = wb.active
        ws.title = title
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        row_number, prev_date = 1, ''

        for item_idx, item in enumerate(items):
            if item_idx == 0:
                for hdr_idx, hdr_name in enumerate(attributes): 
                    ws.cell(row=row_number, column=(hdr_idx + 1)).value = hdr_name
            row_number = row_number + 1
            for attr_idx, attr_name in enumerate(attributes): 
                ws.cell(row=row_number, column=(attr_idx + 1)).value = item[attr_name]

        wb.save(outfile)
        print('file written: {}'.format(outfile))
